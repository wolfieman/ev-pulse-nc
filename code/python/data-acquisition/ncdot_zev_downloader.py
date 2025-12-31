#!/usr/bin/env python3
"""
NCDOT ZEV Registration Data Downloader

Downloads monthly Zero-Emission Vehicle registration data from NC DOT.
Supports flexible date ranges and validates downloaded Excel files.

Usage:
    python ncdot_zev_downloader.py --start 2025-07 --end 2025-10 --outdir ../../../data/raw/ncdot-monthly
    python ncdot_zev_downloader.py --months 2025-07 2025-08 2025-09 2025-10
"""

import argparse
import calendar
import sys
import zipfile
from pathlib import Path
from typing import NamedTuple
from datetime import datetime

import requests

# Optional: openpyxl for Excel validation
try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    InvalidFileException = Exception  # type: ignore[misc,assignment]


class DownloadResult(NamedTuple):
    month: str
    success: bool
    path: Path | None
    message: str


BASE_URL = "https://www.ncdot.gov/initiatives-policies/environmental/climate-change/Documents"
FILENAME_TEMPLATE = "{year}-{month_name}-registration-data.xlsx"
USER_AGENT = {"User-Agent": "NCDOT-ZEV-Downloader/1.0"}
MONTH_NAMES = {i: name.lower() for i, name in enumerate(calendar.month_name) if name}


def parse_month(month_str: str) -> tuple[int, int]:
    """Parse 'YYYY-MM' format to (year, month) tuple."""
    try:
        dt = datetime.strptime(month_str, "%Y-%m")
        return dt.year, dt.month
    except ValueError:
        raise ValueError(f"Invalid month format: {month_str}. Use YYYY-MM (e.g., 2025-07)")


def generate_month_range(start: str, end: str) -> list[tuple[int, int]]:
    """Generate list of (year, month) tuples from start to end inclusive.

    Args:
        start: Start month in YYYY-MM format
        end: End month in YYYY-MM format

    Returns:
        List of (year, month) tuples covering the range

    Raises:
        ValueError: If start month is after end month
    """
    start_year, start_month = parse_month(start)
    end_year, end_month = parse_month(end)

    if (start_year, start_month) > (end_year, end_month):
        raise ValueError(f"Start month ({start}) must be before or equal to end month ({end})")

    months = []
    year, month = start_year, start_month

    while (year, month) <= (end_year, end_month):
        months.append((year, month))
        month += 1
        if month > 12:
            month = 1
            year += 1

    return months


def build_url(year: int, month: int) -> str:
    """Construct download URL for a given year/month."""
    month_name = MONTH_NAMES[month]
    filename = FILENAME_TEMPLATE.format(year=year, month_name=month_name)
    return f"{BASE_URL}/{filename}"


def validate_excel(file_path: Path) -> tuple[bool, str]:
    """Validate that the file is a valid Excel workbook.

    Performs size checks and optionally validates Excel structure using openpyxl.
    The 1KB minimum threshold is based on typical Excel file overhead.

    Args:
        file_path: Path to the Excel file to validate

    Returns:
        Tuple of (is_valid: bool, message: str)
    """
    if not file_path.exists():
        return False, "File does not exist"

    size = file_path.stat().st_size
    if size == 0:
        return False, "File is empty (0 bytes)"

    if size < 1000:  # Excel files have ~1KB overhead minimum
        return False, f"File suspiciously small ({size} bytes)"

    if OPENPYXL_AVAILABLE:
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return True, f"Valid Excel with sheets: {', '.join(sheet_names)}"
        except (InvalidFileException, zipfile.BadZipFile) as e:
            return False, f"Invalid Excel format: {e}"
        except (IOError, OSError) as e:
            return False, f"File read error: {e}"

    return True, f"Downloaded {size:,} bytes (openpyxl not available for deep validation)"


def download_file(url: str, dest: Path, retries: int = 3, timeout: int = 60) -> tuple[bool, str]:
    """Download a file with retry logic.

    Args:
        url: URL to download from
        dest: Destination path for the file
        retries: Number of retry attempts (default: 3)
        timeout: Request timeout in seconds (default: 60)

    Returns:
        Tuple of (success: bool, message: str)
    """
    last_error: str = "Download failed"

    for attempt in range(1, retries + 1):
        response = None
        try:
            response = requests.get(url, headers=USER_AGENT, timeout=timeout, stream=True)
            response.raise_for_status()

            content = response.content
            if not content:
                raise ValueError("Received empty response")

            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(content)

            return True, f"Downloaded {len(content):,} bytes"

        except requests.exceptions.HTTPError as e:
            if response is not None and response.status_code == 404:
                return False, "File not found (404) - data may not be available yet"
            last_error = f"HTTP error: {e}"

        except requests.exceptions.ConnectionError as e:
            last_error = f"Connection error: {e}"

        except requests.exceptions.Timeout as e:
            last_error = f"Request timed out: {e}"

        except requests.exceptions.RequestException as e:
            last_error = f"Request failed: {e}"

        except (IOError, OSError) as e:
            last_error = f"File write error: {e}"

        except ValueError as e:
            last_error = str(e)

    return False, f"{last_error} (after {retries} attempts)"


def download_month(year: int, month: int, outdir: Path) -> DownloadResult:
    """Download and validate data for a specific month."""
    month_label = f"{year}-{month:02d}"
    url = build_url(year, month)
    filename = FILENAME_TEMPLATE.format(year=year, month_name=MONTH_NAMES[month])
    dest = outdir / filename

    # Check if already exists and valid
    if dest.exists():
        valid, msg = validate_excel(dest)
        if valid:
            return DownloadResult(month_label, True, dest, f"Already exists: {msg}")
        else:
            print(f"  [{month_label}] Existing file invalid, re-downloading...")
            dest.unlink()

    # Download
    print(f"  [{month_label}] Downloading from {url}")
    success, msg = download_file(url, dest)

    if not success:
        return DownloadResult(month_label, False, None, msg)

    # Validate
    valid, validation_msg = validate_excel(dest)
    if not valid:
        dest.unlink()  # Remove invalid file
        return DownloadResult(month_label, False, None, f"Download succeeded but validation failed: {validation_msg}")

    return DownloadResult(month_label, True, dest, validation_msg)


def main():
    parser = argparse.ArgumentParser(
        description="Download NCDOT ZEV registration data",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --start 2025-07 --end 2025-10
  %(prog)s --months 2025-07 2025-08 2025-09 2025-10
  %(prog)s --start 2025-01 --end 2025-12 --outdir ../../../data/raw/ncdot-monthly
        """
    )

    parser.add_argument("--start", help="Start month (YYYY-MM format)")
    parser.add_argument("--end", help="End month (YYYY-MM format)")
    parser.add_argument("--months", nargs="+", help="Specific months to download (YYYY-MM format)")
    parser.add_argument("--outdir", default=".", help="Output directory (default: current directory)")

    args = parser.parse_args()

    # Determine which months to download
    if args.months:
        months = [parse_month(m) for m in args.months]
    elif args.start and args.end:
        months = generate_month_range(args.start, args.end)
    else:
        parser.error("Provide either --start/--end or --months")

    outdir = Path(args.outdir).resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"NCDOT ZEV Registration Data Downloader")
    print(f"Output directory: {outdir}")
    print(f"Months to download: {len(months)}")
    print()

    results: list[DownloadResult] = []

    for year, month in months:
        result = download_month(year, month, outdir)
        results.append(result)

        status = "[OK]" if result.success else "[FAIL]"
        print(f"  {status} {result.month}: {result.message}")

    # Summary
    print()
    print("=" * 50)
    successful = [r for r in results if r.success]
    failed = [r for r in results if not r.success]

    print(f"Downloaded: {len(successful)}/{len(results)} files")

    if successful:
        print("\nSuccessful downloads:")
        for r in successful:
            print(f"  {r.path}")

    if failed:
        print("\nFailed downloads:")
        for r in failed:
            print(f"  {r.month}: {r.message}")
        sys.exit(1)


if __name__ == "__main__":
    main()
